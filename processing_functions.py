import html
from datetime import date

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from dateutil.parser import parse
import numpy as np


def is_date(string, fuzzy=False):
    """
    Return whether the string can be interpreted as a date.

    :param string: str, string to check for date
    :param fuzzy: bool, ignore unknown tokens in string if True
    """
    try:
        parse(string, fuzzy=fuzzy)
        return True

    except ValueError:
        return False


def process_long_data(raw_df, long_df, processed_data_file_name):
    """Create and process data in long format

    This function reads in the long format data, if it exists, and
    appends the newly processed long format data to it. If the historical
    long format data has the years of data that are also in the new
    long format data, the newest version will overwrite it. This way,
    the long format data always has the most recently processed version.

    :parma raw_df: Raw OLDC data
    :type raw_df: <Dict(<pd.DataFrame>)>
    :param long_df: Data frame of processed long format data
    :type long_df: <pd.DataFrame>
    :param processed_data_file_name: File name of previously processed data
    :type processed_data_file_name: <str>

    :return: Processed and appended long format grantee data
    :rtype: <pd.DataFrame>

    """

    # Years to include
    years_in_oldc_data = [int(x) for x in raw_df["Screen-1"].Fy.unique()]

    # Check if long format data exists in the current processed file
    sheet_names = pd.ExcelFile(processed_data_file_name).sheet_names
    sheet_index = [is_date(sheet_name) for sheet_name in sheet_names]
    if any(sheet_index):
        # Sheet name of long format data
        last_update = date.fromisoformat(
            [name for i, name in enumerate(sheet_names) if sheet_index[i]][0]
        )

        # read in the historical long data
        historical_long_data = pd.read_excel(
            processed_data_file_name, sheet_name=str(last_update)
        )

        # Overwrite old year's processed data with new
        historical_long_data = historical_long_data[
            ~historical_long_data.Year.isin(years_in_oldc_data)
        ]

        # combine the historical long data with new long data
        # and remove any duplicate rows
        historical_long_data = pd.concat([historical_long_data, long_df])
    else:
        historical_long_data = long_df

    # Light processing
    historical_long_data.Element = historical_long_data.Element.str.upper().replace(
        ["GRANTEE NAME", "GRANTEE_NAME", "GRANTEENAME"], "GRANTEENAME"
    )
    historical_long_data = historical_long_data.drop_duplicates(
        ["Grant Type", "State", "Year", "EIN", "Element"], keep="last"
    )

    return historical_long_data


def read_data(filepath_raw, filepath_crosswalk):
    """Read in relevant data.

    This function reads in the necessary sheets from the given file paths
    and returns each one as a dictionary where each value is a data frame
    corresponding to a given sheet. It includes all the sheets of the raw
    data, the sheets of the previously processed data, and the following
    sheets from the lookup data: lookup, cultspec_subawardee, crosswalk.

    :param filepath_raw: File path to the raw OLDC data (this is what will
        be processed)
    :type filepath_raw: <str>
    :param filepath_crosswalk: File path to the lookup data to be used for
        processing
    :type filepath_crosswalk: <str>

    :return: Data frames corresponding to the given sheets, except for the
        raw data, which is returned as a dictionary of data frames
        corresponding to the relevant sheets
    :rtype: <pd.DataFrame>; raw_data: <Dict<pd.DataFrame>>
    """

    # Read in raw data
    raw_data = pd.read_excel(
        filepath_raw,
        sheet_name=None,
        parse_dates=True,
        dtype={"Grantee Zip4": "str", "Grantee Zip5": "str"},
    )

    print("Reading in raw data - COMPLETE")

    # Load the lookup table
    lookup_data_based = pd.read_excel(filepath_crosswalk, sheet_name="lookup")
    subawardee_lookup = pd.read_excel(
        filepath_crosswalk, sheet_name="cultspec_subawardee"
    )
    print(">>> cultspec headers:", subawardee_lookup.columns.tolist())
    subawardee_lookup = subawardee_lookup.drop_duplicates(subset=["SubAwdCultSpecf"])

    field_names_conversion = pd.read_excel(filepath_crosswalk, "crosswalk")
    print("Reading in the lookup table - COMPLETE")

    return raw_data, lookup_data_based, subawardee_lookup, field_names_conversion


def process_raw_data(raw_df, coalitions = False):
    """Do some light processing on raw data.

    Iterate through each sheet in the raw data and apply light processing.

    :param raw_df: The raw data to be processed
    :type raw_df: <Dict<pd.DataFrame>>

    :return: Lightly processed data
    :rtype: <Dict<pd.DataFrame>>
    """

    # Keep a unique grantee name per EIN. Use screen 1 as reference
    unique_grantee_names = (
        raw_df["Screen-1"].groupby(["RptEin"]
                                   )["GranteeName"].first().reset_index()
    )

    for sheet in raw_df.keys():
        df = raw_df[sheet]

        # Map territories to states
        if not coalitions:
            if "GranteeTypeTxt" in df.columns:
                df.loc[df.GranteeTypeTxt == "Territory",
                    "GranteeTypeTxt"] = "State"

        # Impute NaN values where necessary
        df = df.replace("nan", np.nan)
        df = df.replace("", np.nan)

        if "GranteeName" in df.columns:
            # Use unique grantee name per EIN
            df = df.drop("GranteeName", axis=1).merge(
                unique_grantee_names, how="left", on="RptEin"
            )

        # Trim whitespace
        df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

        raw_df[sheet] = df

    return raw_df


def process_subawardee_data(df, subawardee_lookup, receipt_ids_to_keep):
    """Process subawardee data.

    This function does light processing on the subawardee data and adds
    lookup columns from the subawardee lookup table. It only includes rows
    that are in the receipt_ids_to_keep list.

    :param df: Subawardee data frame to process
    :type df: <pd.DataFrame>
    :param subawardee_lookup: Lookup data frame
    :type subawardee_lookup: <pd.DataFrame>
    :param receipt_ids_to_keep: List of rpt-receipt-ids to filter on
    :type receipt_ids_to_keep: <List<str>>

    :return: Processed and cleaned subawardee data
    :rtype: <pd.DataFrame>
    """

    # Clean up text columns
    subawardee_clean_html = df["Screen-2"]
    subawardee_text_cols = subawardee_clean_html.select_dtypes(include=[
                                                               "object"])
    subawardee_clean_html[subawardee_text_cols.columns] = subawardee_text_cols.applymap(
        lambda x: html.unescape(str(x))
    ).fillna("")

    # Rename columns for FVPSA Funding Type and Primary Services Type
    subawardee_clean_html.rename(
        columns={
            "II Text - FVPSA Funding Type,PPR FVPSA Subawardee - Maze Grid Input Row": "Subawardee - FVPSA Funding Type",
            "II Text - Primary Services Type,PPR FVPSA Subawardee - Maze Grid Input Row": "Subawardee - FVPSA Primary Services Type",
        },
        inplace=True,
    )

    # Use lookup to get predefined categories
    subawardee_processed = subawardee_clean_html.merge(
        subawardee_lookup,
        left_on="Subawardee List - Underserved or culturally- and linguistically-specific population",
        right_on="SubAwdCultSpecf",
        how="left",
    )

    # Resolve NA values
    subawardee_processed = subawardee_processed.replace("nan", np.nan)

    # Filter records by Rpt-Receipt-Id (should only include subawardees that map to grantees in processed grantee data)
    final_subawardee = subawardee_processed[
        subawardee_processed["Rpt-Receipt-Id"].isin(receipt_ids_to_keep)
    ]

    return final_subawardee


def get_removable_cols(lookup_data, long_data):
    """Get the columns to filter on.

    This function gets the list of columns that can be filtered on to remove
    empty values, based on the long data. It maps the Meta Name Description
    to the variable column in the long data to find the set difference.

    :param lookup_data: Lookup sheet from lookup data
    :type lookup_data: <pd.DataFrame>
    :param long_data: Long data frame of processed data
    :type long_data: <pd.DataFrame>

    :return: List of field names
    :rtype: <List<str>>
    """

    lookup_set = set(lookup_data["Meta Name Description"].str.upper())
    lookup_set.add("SUBAWARDEE_SHELTER_TOTAL")  # Subawardee - Shelter Total
    # Subawardee - Non-Shelter Total
    lookup_set.add("SUBAWARDEE_NONSHELTER_TOTAL")
    original_set = {str(x).upper() for x in set(long_data.variable)}
    removable_cols = [x for x in (set.difference(original_set, lookup_set))]

    return removable_cols


def save_to_final_workbook(df_to_save, sheet_name, historical_workbook=None):
    """Save sheet to workbook.

    This function saves a given data frame as a sheet to a workbook. If
    no workbook exists, create one and title it as the given sheet name.

    :param df_to_save: Data frame to save to the workbook
    :type df_to_save: <pd.DataFrame>
    :param sheet_name: Name of sheet to save data frame to
    :type sheet_name: <str>
    :param historical_workbook: Workbook to save to. Default is None
    :type historical_workbook: <openpyxl.Workbook>

    :return: Workbook with new sheet saved
    :rtype: <openpyxl.Workbook>
    """

    # If no workbook is given, create one and title it as sheet_name
    if historical_workbook is None:
        historical_workbook = Workbook()
        ws = historical_workbook.active
        ws.title = sheet_name
    else:
        # Create sheet
        ws = historical_workbook.create_sheet(sheet_name)

    this_shape = df_to_save.shape

    # Save
    for r in dataframe_to_rows(df_to_save, index=False, header=True):
        ws.append(r)

    # Add protection
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.enable()
    ws.auto_filter.ref = f"A1:{get_column_letter(this_shape[1])}{this_shape[0]}"

    return historical_workbook


def parse_ein(old):
    """Parse the EIN.

    Remove the brackets and spaces from the EIN number for easier parsing
    and comparison. Example: EIN from [1 236003113 A1] to 1236003113A1

    :param old: The EIN to parse
    :type old: <str>

    :return: The parsed EIN
    :rtype: <str>
    """

    return old[1:2] + old[3:-4] + old[-3:-1]


def lookup_name_from_ein(EIN, df):
    """Get the grantee name from the EIN.

    This function returns the grantee name from the most recent submission
    for a given EIN. If there is no data for that EIN, return None.

    :param EIN: Grantee EIN
    :type EIN: <str>
    :param df: Data frame to filter on
    :type df: <pd.DataFrame>

    :return: The name of the most recent year's submission, or None
    :rtype: <str>
    """

    # Find all years that this grantee had data
    available_years = df.query("EIN == @EIN").Fy.unique()

    # If this grantee had data, return the name of the most recent year's submission
    # Otherwise, return None
    if len(available_years) > 0:
        best_year = max(available_years)
        name = (
            df.query("EIN == @EIN and Fy == @best_year").GranteeName.iloc[0]
            if best_year >= 2018
            else None
        )
        return html.unescape(name)
    else:
        return None


def create_metadata_sheet(
    workbook,
    wide_data,
    all_states,
    include_states=True,
    include_tribes=True,
    sheet_location=-1,
):
    """Create Meta Data Sheet.

    This function creates a sheet with metadata information, including the
    number of states & tribes reporting each year, the timestamp of the last
    data processing, and the list of missing states for each year.

    :param workbook: Workbook to save sheet to
    :type workbook: <openpyxl.Workbook>
    :param wide_data: Data to derive metadata from
    :type wide_data: <pd.DataFrame>
    :param all_states: List of states to include
    :type all_states: <List<str>>
    :param include_states: Boolean, to include state info in metadata or not
    :type include_states: <Bool>, default is True
    :param include_tribes: Boolean, to include tribe info in metadata or not
    :type include_tribes: <Bool>, default is True
    :param sheet_location: Location of sheet in workbook
    :type sheet_location: <Int>, default is -1

    :return: Workbook with meta data sheet
    :rtype: <openpyxl.Workbook>
    """

    # Create sheet in workbook
    if sheet_location != -1:
        ws = workbook.create_sheet("Metadata", sheet_location)
    else:
        ws = workbook.create_sheet("Metadata")

    # Assign last processing date to today
    ws["A1"].value = "Last data processing:"
    ws["A2"].value = str(date.today())

    # Create metadata
    current_rowindex = 2
    ws["C1"].value = "Year"
    if include_states:
        ws[f"C{current_rowindex}"].value = "Number of states reporting"
        current_rowindex += 1
    if include_tribes:
        ws[f"C{current_rowindex}"].value = "Number of tribes reporting"
        current_rowindex += 1
    if include_states:
        ws[f"C{current_rowindex}"].value = "List of missing states"
        current_rowindex += 1

    # Years to report from processed data
    available_years = sorted(wide_data.Year.unique())

    # Create a list of the number of states and tribes and how many are missing for each year
    for year_i, year in enumerate(available_years):
        if include_states:
            states_present = wide_data.query(
                "Year == @year and `Grant Type` == 'State'"
            ).State.unique()
            missing_states = sorted(
                list(set.difference(set(all_states), set(states_present)))
            )

        ws[f"{get_column_letter(year_i + 4)}1"].value = year
        current_rowindex = 2

        if include_states:
            ws[f"{get_column_letter(year_i + 4)}{current_rowindex}"].value = len(
                wide_data.query(
                    "Year == @year and `Grant Type` == 'State'"
                ).EIN.unique()
            )
            current_rowindex += 1
        if include_tribes:
            ws[f"{get_column_letter(year_i + 4)}{current_rowindex}"].value = len(
                wide_data.query(
                    "Year == @year and `Grant Type` == 'Tribe'"
                ).EIN.unique()
            )
            current_rowindex += 1
        if include_states:
            for state_i, state in enumerate(missing_states):
                ws[
                    f"{get_column_letter(year_i + 4)}{current_rowindex + state_i}"
                ].value = state

    return workbook


def calculate_gender_totals(df, genders):
    """Calculate total clients served.

    To calculate the total number of clients served, first iterate over the
    shelter and non-shelter total clients served by gender, and then add
    those totals over all genders. i.e. Shelter total = # women shelter
    clients + # men shelter clients + # children shelter clients + # shelter
    Not Specified. Currently, 0 totals are converted to NaN values.

    :param df: Data frame to store shelter and non-shelter totals
    :type df: <pd.DataFrame>
    :param genders: List of genders to sum over, as strings
    :type genders: <list<str>>

    :return: Data frame with shelter and non-shelter totals
    :rtype: <pd.DataFrame>
    """

    # Initialize totals
    df["Shelter Total"] = 0
    df["Non-shelter Total"] = 0

    # Get totals for each gender listed
    for gender in genders:
        # Column headers to reference
        shelter_col = f"Shelter {gender}"
        nonshelter_col = f"Non-shelter {gender}"

        # Initialize to 0
        df.loc[:, gender] = 0

        # Total clients served for gender = shelter + nonshelter totals
        df.loc[:, gender] = df.loc[:, [shelter_col, nonshelter_col]].apply(
            lambda x: (x[0] if pd.notnull(x[0]) else 0)
            + (x[1] if pd.notnull(x[1]) else 0),
            axis=1,
        )

        # Convert 0 counts to NaN
        df.loc[df[gender] == 0, gender] = np.nan

        # Total shelter clients served is the sum of shelter totals for all genders
        df.loc[pd.notna(df[shelter_col]), "Shelter Total"] += df.loc[
            pd.notna(df[shelter_col]), shelter_col
        ]

        # Total non-shelter clients served is the sum of shelter totals for all genders
        df.loc[pd.notna(df[nonshelter_col]), "Non-shelter Total"] += df.loc[
            pd.notna(df[nonshelter_col]), nonshelter_col
        ]

    # Convert 0 shelter and non-shelter totals to NaN
    df.loc[df["Shelter Total"] == 0, "Shelter Total"] = np.nan
    df.loc[df["Non-shelter Total"] == 0, "Non-shelter Total"] = np.nan

    return df


def calculate_total_funds(subawardee_df, state_df, cols_to_merge):
    """Calculate total subawardee funds by state.

    Calculate the total funding amount by adding subawardee funding amounts
    for each state and year. Shelter and nonshelter columns are renamed to
    SUBAWARDEE_SHELTER_TOTAL and SUBAWARDEE_NONSHELTER_TOTAL. This function
    appends the final subawardee shelter type funding totals to the state
    grantee data.

    :param subawardee_df: Data frame of processed subawardee data
    :type subawardee_df: <pd.DataFrame>
    :param state_df: Data frame of processed state grantee data
    :type state_df: <pd.DataFrame>
    :param cols_to_merge: List of column names to merge data frames on
    :type cols_to_merge: List of <str>

    :return: Data frame of state grantee data with total subawardee funding
        amounts appended
    :rtype: <pd.DataFrame>
    """

    # Only states fill out the subawardee portion of the PPR
    new_subawardee_df = subawardee_df.copy(
        deep=True).query("GranteeTypeTxt == 'State'")

    # Standardize the Shelter Type field
    shelter_index = (
        new_subawardee_df["Subawardee List - Type of Subawardee"].str.upper()
        == "SHELTER"
    )
    new_subawardee_df["ShelterType"] = np.where(
        shelter_index, "Shelter", "Non-Shelter")

    # Join state subawardee data to state grantee data
    new_states_processed = state_df.merge(
        new_subawardee_df, how="left", on=cols_to_merge
    )

    # Create total shelter and nonshelter funds for each year and state:
    shelter_compare = (
        new_subawardee_df.groupby(["Fy", "PostalCode", "ProgAcronym", "ShelterType"])[
            "Subawardee List - FVPSA Funding Amount"
        ]
        .sum()
        .unstack()
        .reset_index()
        .fillna(0)
        .rename(
            columns={
                "Non-Shelter": "SUBAWARDEE_NONSHELTER_TOTAL",
                "Shelter": "SUBAWARDEE_SHELTER_TOTAL",
            }
        )
    )

    # Merge total shelter/non-shelter funding amounts back onto states data frame
    new_states_processed = new_states_processed.merge(
        shelter_compare, how="left", on=["Fy", "PostalCode", "ProgAcronym"]
    )

    return new_states_processed


def join_on_meta_name_desc(long_data, meta_name_df, year=None):
    """Join to lookup table.

    This function takes the long format processed grantee data and merges it
    to the lookup table. The lookup table contains the relevant columns to
    match on, and the Element column, which can be used to merge onto the
    final, cleaned column names. This function also removes empty values and
    subsets the merged data frame to the relevant fields.

    :param long_data: Data frame of grantee data, in long format, with
        the "variable" column included
    :type long_data: <pd.DataFrame>
    :param meta_name_df: Data frame of the lookup sheet that contains the
        "Meta Name Description" column, to be merged on the long_data
    :type meta_name_df: <pd.DataFrame>

    :return: Data frame of merged data, with empty values removed, a new
        Element column, and subset to only the relevant columns
    :rtype: <pd.DataFrame>
    """

    print("Unique variables in long_data before merge:", long_data["variable"].unique())
    if year == 2024:
        # Only apply this mapping in 2024
        question_mapping = {
            "H-02 What does the FVPSA grant allow you to do that you wouldn¿t be able to do without this funding?":
                "H-02 WHAT DOES THE FVPSA GRANT ALLOW YOU TO DO THAT YOU WOULDN¿T BE ABLE TO DO WITHOUT THIS FUNDING?...49",
            "H-02 What does the FVPSA grant allow you to do that you wouldn¿t be able to do without this funding?.1":
                "H-02 WHAT DOES THE FVPSA GRANT ALLOW YOU TO DO THAT YOU WOULDN¿T BE ABLE TO DO WITHOUT THIS FUNDING?...50",
        }
        long_data['variable'] = long_data['variable'].replace(question_mapping)


    # Join the long data on the lookup table meta name description
    long_data.variable = long_data.variable.str.upper()
    all_long_data = long_data.merge(
        meta_name_df,
        how="left",
        left_on=["variable"],
        right_on=["Meta Name Description"],
    ).drop(columns="Meta Name Description")


    # Grab columns that are in element (from lookup table), but not in variable (from long data)
    # This only grabs the relevant variables to include based on the lookup table
    removable_cols = get_removable_cols(meta_name_df, all_long_data)


    # Add the engineered SUBAWARDEE_SHELTER_TOTAL columns to the Element column (not currently included in lookup table)
    subawardee_shelter_index = all_long_data.variable.isin(
        ["SUBAWARDEE_SHELTER_TOTAL", "SUBAWARDEE_NONSHELTER_TOTAL"]
    )
    all_long_data.loc[subawardee_shelter_index, "Element"] = all_long_data.loc[
        subawardee_shelter_index, "variable"
    ]

    # Drop empty values and clean up column names
    all_long_data = (
        all_long_data[~all_long_data.variable.str.upper().isin(removable_cols)]
        .dropna(subset=["Element"])
        .drop(columns=["variable"])
    )

    all_long_data = all_long_data[
        [
            "GranteeTypeTxt",
            "Fy",
            "ProgAcronym",
            "PostalCode",
            "EIN",
            "Clients",
            "In Use",
            "Demo",
            "TypeService",
            "Outcomes",
            "Element",
            "value",
        ]
    ].rename(
        columns={
            "value": "Value",
            "GranteeTypeTxt": "Grant Type",
            "PostalCode": "State",
            "Fy": "Year",
            "ProgAcronym": "Program Acronym",
        }
    )

    return all_long_data


def service_outcome_survey_type_helper(
        outcome_dat, xw_dat, outcome_columns, id_cols, survey_type_ind, survey_type_str, safety=False):
    """Transform Service Outcome Data by Service Type

    This function takes the service outcome grantee data, which is referenced
    in the crosswalk file, for a particular service type and converts it to 
    long format and cleans up the column names. 

    :param outcome_dat: Data frame of grantee service outcome data
    :type outcome_dat: <pd.DataFrame>
    :param xw_dat: Data frame of the lookup sheet that contains the
        "Meta Name Description" column, to be merged on the outcome_dat
    :type xw_dat: <pd.DataFrame>
    :param outcome_columns: Service outcome columns to subset on
    :type outcome_columns: <List(<str>)>
    :param id_cols: Unique identification columns in outcome_dat
    :type id_cols: <List(<str>)>
    :param survey_type_ind: The survey type indicator from the raw columns
    :type survey_type_ind: <str>
    :param survey_type_str: The name of the survey type to assign the new column
    :type survey_type_str: <str>
    :param safety: Whether this is a safety planning survey type or not
    :type safety: <bool>

    :return: Survey type and value in long format
    :rtype: <pd.DataFrame>
    """

    # Subset appropriate column names for this service type
    if not safety:
        survey_type_cols = [
            c for c in outcome_columns if survey_type_ind in c and "SAFTY" not in c and "SAFETY" not in c]
    else:
        survey_type_cols = [
            c for c in outcome_columns if survey_type_ind in c and ("SAFTY" in c or "SAFETY" in c)]

    # Subset data
    survey_type = outcome_dat[id_cols + survey_type_cols]

    # Transform to long format and make names pretty
    survey_type_dat = pd.melt(
        survey_type,
        value_vars=survey_type_cols,
        id_vars=["GRANTEETYPETXT",
                 "FY",
                 "PROGACRONYM",
                 "POSTALCODE",
                 "EIN"],
        var_name="Survey Type",
        value_name=survey_type_str
    ).merge(xw_dat, how="left", left_on=["Survey Type"], right_on=["Meta Name Description"])[
        id_cols + [
            "Label",
            survey_type_str
        ]
    ].\
        rename(
        columns={
            "Label": "Survey Type"
        }
    )

    # Standardize survey type values
    survey_type_dat["Survey Type"] = [
        "Counseling Survey" if "Counseling" in s
        else "Shelter Survey" if "Shelter" in s
        else "Support Group Survey" if "Support Group" in s
        else "Support Services and Advocacy Survey" if "Support Services and Advocacy" in s
        else "Total" for s in survey_type_dat["Survey Type"]]

    # Sort
    survey_type_dat.sort_values(
        by=id_cols+["Survey Type"], inplace=True)

    return(survey_type_dat.reset_index(drop=True))


def service_outcome_transform(processed_dat, xw_dat):
    """Transform Service Outcome Data

    This function takes the service outcome grantee data, which is referenced
    in the crosswalk file, and transforms it to resemble the format in the 
    PPR. 

    :param processed_dat: Processed raw grantee data
    :type processed_dat: <pd.DataFrame>
    :param xw_dat: Data frame of the lookup sheet that contains the
        "Meta Name Description" column, to be merged on the outcome_dat
    :type xw_dat: <pd.DataFrame>

    :return: Service outcome data in long format
    :rtype: <pd.DataFrame>
    """

    dat = processed_dat.copy()

    # Identify service outcome columns
    dat.columns = map(str.upper, dat.columns)
    outcome_columns = list(xw_dat.loc[xw_dat.Group_Description ==
                                      "Service Outcome", "Meta Name Description"])
    # Unique identifier columns
    id_cols = ["GRANTEETYPETXT",
               "FY",
               "PROGACRONYM",
               "POSTALCODE",
               "EIN"]

    # Subset
    outcome_dat = dat[
        id_cols +
        outcome_columns
    ]

    # Transform service data for each service type
    survey_type_map = {
        "Number of Surveys Completed Resource Outcome": {"ind": "NUMBER OF SURVEYS", "safety": False},
        "Number of Yes Responses to Resource Outcome": {"ind": "NUMBER OF YES RESPONSES", "safety": False},
        "Percent Responses Resource Outcome": {"ind": "PERCENTAGE", "safety": False},
        "Number of Surveys Completed Safety Planning": {"ind": "NUMBER OF SURVEYS", "safety": True},
        "Number of Yes Responses to Safety Planning": {"ind": "NUMBER OF YES RESPONSES", "safety": True},
        "Percent Responses Safety Planning": {"ind": "PERCENTAGE", "safety": True}
    }

    # Initialize output
    outcome_dat_final = dat[id_cols]
    for surv in range(0, len(survey_type_map)):

        # Get survey type name
        this_str = list(survey_type_map.keys())[surv]

        # Transform survey type
        surv_dat = service_outcome_survey_type_helper(outcome_dat=outcome_dat,
                                                      xw_dat=xw_dat,
                                                      outcome_columns=outcome_columns,
                                                      id_cols=id_cols,
                                                      survey_type_ind=survey_type_map[this_str]["ind"],
                                                      survey_type_str=this_str,
                                                      safety=survey_type_map[this_str]["safety"])

        if "Survey Type" in outcome_dat_final.columns:
            join_columns = id_cols + ["Survey Type"]
        else:
            join_columns = id_cols

        # Concat
        outcome_dat_final = outcome_dat_final.merge(
            surv_dat,
            how="inner",
            on=join_columns
        )

    # Make sure identifier columns are readable
    outcome_dat_final = outcome_dat_final.\
        rename(
            columns={
                "GRANTEETYPETXT": "Grant Type",
                "FY": "Year",
                "PROGACRONYM": "Program Acronym",
                "POSTALCODE": "State"
            }
        )

    return(outcome_dat_final)


def create_codetxt_table(processed_data):
    """Create Codetext Table for Metadata Sheet.

    From the raw processed grantee data:
        1) Get the number of entries in the data for each year, grantee type,
            code text, and program
        2) Sort by year (desc), grantee type, and code text
        3) Pivot to wide format

    :param processed_data: Data frame of raw processed grantee data
    :type processed_data: <pd.DataFrame>

    :return: Data frame of summarized meta data
    :rtype: <pd.DataFrame>
    """

    codetxt_table = (
        processed_data.groupby(
            ["Fy", "GranteeTypeTxt", "CodeTxt", "ProgAcronym"])
        .size()
        .to_frame("count")
        .reindex(
            pd.MultiIndex.from_product(
                [
                    processed_data["Fy"].unique(),
                    processed_data["GranteeTypeTxt"].unique(),
                    processed_data["CodeTxt"].unique(),
                    processed_data["ProgAcronym"].unique(),
                ]
            ),
            fill_value=0,
        )
        .reset_index()
        .sort_values(by=["level_0", "level_1", "level_2"])
        .pivot(
            index="level_2", columns=["level_0", "level_1", "level_3"], values="count"
        )
        .rename_axis("CodeTxt")
        .reset_index(level=0)
        .transpose()
        .reset_index()
        .transpose()
    )

    return codetxt_table


def replace_duplicate_columns(df, replacements=None):
    """Replace duplicated column names.

    This function finds the columns in df that contain the substrings
    provided in duplicate_col_list and replaces them with the columns
    provided in replacements. For example, there are two columns with the
    substring 'H-02 What does the FVPSA grant allow you to do that you
    wouldn¿t be able to do without this funding?' in the OLDC data, one of
    them with a '.1' appended to it. This should be replaced with the proper
    name. This is generic enough that it can be used on future data that may
    or may not contain different duplicates. This function assumes that the
    substring is the column containing the first and correct column name,
    and all other columns following it that contain that substring need to
    be renamed.

    :param df: Data frame to find and replace column names from
    :type df: <pd.DataFrame>
    :param replacements: Dictionary where the keys are the column name
        substrings that contain duplicates in the df and the values are
        their replacements.
    :type replacements: <Dict<str>: <str>>

    :return: Data frame with renamed columns
    :rtype: <pd.DataFrame>
    """

    # The list of column name substrings that are duplicated in df
    duplicate_col_list = replacements.keys()

    # If None, there's nothing to rename
    if duplicate_col_list is None:
        return

    # Rename columns for each substring
    for dup_col_substring in duplicate_col_list:
        # List of duplicated columns based on substring
        duplicate_cols = [
            col
            for col in df.columns
            if dup_col_substring in col and col != dup_col_substring
        ]

        # Only rename if there are duplicated columns
        if len(duplicate_cols) > 0:
            dup_col_replace = replacements[dup_col_substring]

            # Input error could cause the number of replacements to vary from the number of duplicates
            if len(dup_col_replace) != len(duplicate_cols):
                Warning(
                    f"The number of duplicate columns that exist does not match the number of replacements. Only "
                    f"replacing the first {len(dup_col_replace)} instances..."
                )

            # Rename all the duplicate columns
            for dup_col, replace in zip(duplicate_cols, dup_col_replace):
                df = df.rename(columns={dup_col: replace})

    return df
